#module importing
from MODULE_TEST import *

#for csv files operation
import csv
from openpyxl import load_workbook



#-----------extracting dataset------------------

def csv_datasheet():
    wb = load_workbook(filename='CONTENT  SHEET__GYAAN CONNECT.xlsx')

    sheet = wb.active

    csv_data = []

    for value in sheet.iter_rows(values_only = True):
        csv_data.append(list(value))

    with open("datasheet.csv",'w') as csv_obj:
        writer = csv.writer(csv_obj, delimiter = ',')
        for line in csv_data:
            writer.writerow(line)
        csv_obj.close()
    return csv_data


def create_csv(filename,userid):
    
    with open(f"user{id}.csv",'w') as user_preference_file:
        writer = csv.writer(user_preference_file,delimiter=',')







        
#-----------------function for dataset loaded for comparision----------------------



def fetching_dataset(csv_file):
    data_sheet = {"subject":[],"channel_link":[],"rec_link":[],"expertise":[],"lang":[],"course":[],"video_level":[]}

    for link in csv_file():

        data_sheet["subject"].append(link[0]) 
        data_sheet["channel_link"].append(link[1]) 
        data_sheet["rec_link"].append(link[2]) 
        data_sheet["expertise"].append(link[3]) 
        data_sheet["lang"].append(link[4]) 
        data_sheet["course"].append(link[5]) 
        data_sheet["video_level"].append(link[6]) 
    
    return data_sheet

#--------------------creating the stats for the playlist-----------------------------

subject_choice = input("What do you want to study? e.g Python, Java, C++, Maths etc - ").lower()
dic = {subject_choice: []}

for count, subject in enumerate(fetching_dataset(csv_datasheet)["subject"]):
    if subject_choice in subject.lower():
        dic[subject_choice].append(fetching_dataset(csv_datasheet)["rec_link"][count])

for key in dic.values():
    
    pass



    

